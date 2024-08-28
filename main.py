import csv
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

class PPR:
    weights = {
        "PASS_YDS": 0.04,
        "PASS_TDS": 4,
        "PASS_INTS": -2,
        "RUSH_YDS": 0.1,
        "RUSH_TDS": 6,
        "REC_YDS": 0.1,
        "REC": 1,
        "REC_TDS": 6,
        "FL": -2,
        # All stats appearing in data source should be accounted for, for completeness
        "PASS_ATT": 0,
        "PASS_CMP": 0,
        "RUSH_ATT": 0
    }

    @staticmethod
    def score(stat_dict):
        if stat_dict == {}:
            return None
        score = 0
        for stat in stat_dict:
            score += stat_dict[stat] * PPR.weights[stat]
        # Divide by 17 to get per game score as it is easier to visualize
        return score / 17


class Player:
    def __init__(self, name, position):
        self.name = name
        self.position = position
        self.average_projection = {}
        self.high_projection = {}
        self.low_projection = {}

    def get_score(self, scoring=PPR):
        return (scoring.score(self.low_projection),
                scoring.score(self.average_projection),
                scoring.score(self.high_projection))

    def set_projections(self, projection_type, data):
        if projection_type == 'average':
            self.average_projection = data
        elif projection_type == 'high':
            self.high_projection = data
        elif projection_type == 'low':
            self.low_projection = data

    def __repr__(self):
        return f"Player({self.position} {self.name})"


def parse_csv(file_path, position):
    players = {}
    current_player = None
    current_type = None

    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        headers = next(reader)  # Read the header

        # First row after headers is garbage
        next(reader)

        # Determine the columns to include (ignore "Team" and "FPTS")
        relevant_columns = [i for i, header in enumerate(headers) if header not in ("Team", "FPTS")]

        for row in reader:
            # Some rows at the bottom are empty
            if len(row) == 1:
                continue
            if row[0].strip():  # New player entry
                player_name = row[0].strip()
                current_player = Player(player_name, position)
                current_type = 'average'
                players[player_name] = current_player

            elif 'high' in row[1]:
                current_type = 'high'
            elif 'low' in row[1]:
                current_type = 'low'

            # Parse stats into a dictionary, skipping "Team" and "FPTS"
            stats = {}
            for i in relevant_columns[1:]:  # Skip the first relevant column (Player)
                if row[i].strip():
                    stats[headers[i]] = float(row[i].replace(',',''))

            current_player.set_projections(current_type, stats)

    return players


if __name__ == '__main__':
    # Gather projections from Fantasy Pros
    data_dir = "fp_data"
    players = {"QB": {}, "TE": {}, "RB": {}, "WR": {}, "All": {}, "Flex": {}}
    for f in os.listdir(data_dir):
        if f.endswith(".csv"):
            position = f.split(".")[0].split("_")[-1]
            player_dict = parse_csv(os.path.join(data_dir, f), position)
            players[position] = player_dict
            players["All"].update(player_dict)
            if position != "QB":
                players["Flex"].update(player_dict)

    # Baseline is the number of that position drafted through 9 round last year plus 1
    baseline = {
        "QB": 11,
        "TE": 12,
        "RB": 39,
        "WR": 49
    }
    baseline["All"] = sum(baseline.values())
    baseline["Flex"] = sum(baseline.values()) - baseline["QB"]
    # Only keep a certain amount of player projections per position
    # Rule of thumb here will be double the amount needed for rosters
    position_limits = {
        "QB": 12 * 2,
        "TE": 12 * 2,
        "RB": 36 * 2,
        "WR": 48 * 2
    }
    position_limits["All"] = sum(position_limits.values())
    position_limits["Flex"] = sum(position_limits.values()) - position_limits["QB"]
    # Find projection of baseline player at every position
    # This is what we will compare against
    data = {}  # player_name : positional_values; to use for All sheet
    baseline_players = {}
    for position in baseline.keys():
        sorted_players = sorted(players[position].values(), key=lambda P: P.get_score()[1], reverse=True)
        baseline_player = sorted_players[baseline[position]]
        baseline_avg = baseline_player.get_score()[1]
        baseline_players[position] = baseline_player

        # Map value of each player relative to the baseline player, store in CSV for output
        with open(os.path.join("output", f"{position}.csv"), "w") as f:
            f.write("Name,Low,Avg,High,Range\n" if position not in ["All", "Flex"]
                    else "Name,Position,Low,Avg,High,Range,Pos. Low,Pos. Avg,Pos. High\n")
            for player in sorted_players[:position_limits[position]]:
                low, avg, high = player.get_score()
                if position not in ["All", "Flex"]:
                    row = [player.name, f"{low - baseline_avg:.1f}", f"{avg - baseline_avg:.1f}",
                           f"{high - baseline_avg:.1f}", f"{high - low:.1f}"]
                    f.write(','.join(row) + '\n')
                    row[0] = position
                    row = row[:-1]
                    data[player.name] = row
                else:
                    # There is a chance a player in the "All" sheet was not good enough
                    # to be on their own position sheet
                    if player.name not in data:
                        positional_baseline_player = baseline_players[player.position]
                        positional_baseline_avg = positional_baseline_player.get_score()[1]
                        row = [player.position, f"{low - positional_baseline_avg:.1f}",
                               f"{avg - positional_baseline_avg:.1f}", f"{high - positional_baseline_avg:.1f}"]
                        data[player.name] = row
                    row = [player.name, data[player.name][0], f"{low - baseline_avg:.1f}", f"{avg - baseline_avg:.1f}",
                           f"{high - baseline_avg:.1f}", f"{high - low:.1f}"] + data[player.name][1:]
                    f.write(','.join(row) + '\n')

    # Convert CSV output into formatted Excel files
    # Create a new Excel workbook
    wb = Workbook()

    for file_name in sorted(os.listdir("output")):
        if not file_name.endswith(".csv"):
            continue
        # Load each CSV into a pandas DataFrame
        df = pd.read_csv(os.path.join("output", file_name))

        # Add a new sheet to the workbook
        ws = wb.create_sheet(title=file_name.split(".")[0])

        # Write the DataFrame to the new worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Set the width of the "Name" column to fit the longest name
        if "Name" in df.columns:
            max_name_length = df["Name"].str.len().max()
            ws.column_dimensions['A'].width = max_name_length + 2  # Adjust width with some padding

        # Apply a heatmap to the all columns except "Name" and "Position"
        for column_name in [col for col in df.columns if col not in ["Name", "Position"]]:
            if column_name in df.columns:
                col_index = df.columns.get_loc(column_name) + 1  # +1 because openpyxl is 1-indexed

                if column_name == "Range":
                    # Define the color scale as a gradient over percentiles, lower range is better
                    heatmap_rule = ColorScaleRule(
                        start_type='percentile', start_value=0, start_color='00FF00',  # Green for low values
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF',  # White for middle values
                        end_type='percentile', end_value=100, end_color='FF6347'  # Darker red for high values
                    )
                else:
                    # Define the color scale with 0 as the midpoint
                    heatmap_rule = ColorScaleRule(
                        start_type='min', start_value=None, start_color='FF6347',  # Darker red for minimum values
                        mid_type='num', mid_value=0, mid_color='FFFFFF',  # White for 0
                        end_type='max', end_value=None, end_color='00FF00'  # Green for maximum values
                    )

                # Apply the rule to the appropriate column
                col_letter = chr(64 + col_index)  # Convert column index to Excel column letter
                ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', heatmap_rule)

    # Remove the default first sheet created with the workbook (if it hasn't been used)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save the workbook
    wb.save("DraftSheet.xlsx")

    print("Done")


